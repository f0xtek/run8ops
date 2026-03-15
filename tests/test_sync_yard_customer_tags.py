import os
import pytest
import pandas as pd
from openpyxl import load_workbook

from sync_yard_customer_tags import parse_table, save_sheet

BASE_URL = "https://www.example.com"


def make_html(rows: list[list[str]], headers: list[str] | None = None, image_href: str | None = None) -> bytes:
    """Build minimal table HTML for use in tests.

    rows: each row includes the leading row-ID cell as its first element,
          mirroring the real page structure before the first-column is dropped.
    """
    th_row = ""
    if headers:
        ths = "".join(f"<th>{h}</th>" for h in headers)
        th_row = f"<tr>{ths}</tr>"

    tr_rows = []
    for row in rows:
        cells = []
        for i, cell in enumerate(row):
            if i == 1 and image_href:
                cells.append(f'<td><a href="{image_href}">img</a></td>')
            else:
                cells.append(f"<td>{cell}</td>")
        tr_rows.append(f"<tr>{''.join(cells)}</tr>")

    html = f"<table>{th_row}{''.join(tr_rows)}</table>"
    return html.encode()


# ---------------------------------------------------------------------------
# parse_table
# ---------------------------------------------------------------------------


def test_parse_table_returns_dataframe():
    html = make_html([["1", "", "11B", "Acme", "BNSF"]])
    df = parse_table(html, BASE_URL)
    assert isinstance(df, pd.DataFrame)


def test_parse_table_row_count():
    html = make_html([["1", "", "11B", "Acme", "BNSF"], ["2", "", "22C", "Beta", "UP"]])
    df = parse_table(html, BASE_URL)
    assert len(df) == 2


def test_parse_table_drops_first_column():
    # Row ID "1" in the first cell should not appear anywhere in the output
    html = make_html([["1", "", "11B", "Acme"]])
    df = parse_table(html, BASE_URL)
    assert "1" not in df.values


def test_parse_table_first_column_dropped_reduces_column_count():
    # 5-cell rows → 4 output columns (one dropped)
    html = make_html([["1", "", "11B", "Acme", "BNSF"]])
    df = parse_table(html, BASE_URL)
    assert len(df.columns) == 4


def test_parse_table_image_column_has_heading():
    html = make_html(rows=[["1", "", "11B"]], headers=["Tag", "Railroad"])
    df = parse_table(html, BASE_URL)
    assert df.columns[0] == "Image URL"


def test_parse_table_headers_assigned():
    # 4-cell rows, 3 <th> tags — after dropping col 0: 3 cols, 3 headers, no padding
    # image column header is overridden to "Image URL"
    html = make_html(
        rows=[["1", "", "11B", "Acme"]],
        headers=["Tag", "Customer", "Railroad"],
    )
    df = parse_table(html, BASE_URL)
    assert list(df.columns) == ["Image URL", "Customer", "Railroad"]


def test_parse_table_headers_padded_when_fewer_than_columns():
    # 4-cell rows, 2 <th> tags — after dropping col 0: 3 cols, pad 1 left, then set image heading
    html = make_html(rows=[["1", "", "11B", "Acme"]], headers=["Customer", "Railroad"])
    df = parse_table(html, BASE_URL)
    assert len(df.columns) == 3
    assert df.columns[0] == "Image URL"
    assert df.columns[1] == "Customer"
    assert df.columns[2] == "Railroad"


def test_parse_table_no_headers_uses_default_index():
    # 3-cell rows → drop col 0 → 2 output columns → default integer index
    html = make_html(rows=[["1", "", "11B"]])
    df = parse_table(html, BASE_URL)
    assert list(df.columns) == [0, 1]


def test_parse_table_image_url_resolved_to_absolute():
    html = make_html(
        rows=[["1", "", "11B"]],
        image_href="/reference/destinationtags/customers/socal/11B.jpg",
    )
    df = parse_table(html, BASE_URL)
    assert df.iloc[0, 0] == "https://www.example.com/reference/destinationtags/customers/socal/11B.jpg"


def test_parse_table_missing_image_is_empty_string():
    html = make_html(rows=[["1", "", "11B"]])
    df = parse_table(html, BASE_URL)
    assert df.iloc[0, 0] == ""


def test_parse_table_no_table_raises():
    html = b"<html><body><p>No table here</p></body></html>"
    with pytest.raises(ValueError, match="No table found"):
        parse_table(html, BASE_URL)


def test_parse_table_empty_rows_raises():
    html = b"<table><tr><th>Col</th></tr></table>"
    with pytest.raises(ValueError, match="no data rows"):
        parse_table(html, BASE_URL)


# ---------------------------------------------------------------------------
# save_sheet
# ---------------------------------------------------------------------------


def test_save_sheet_creates_new_file(tmp_path):
    output = str(tmp_path / "out.xlsx")
    df = pd.DataFrame([["a", "b"]], columns=["X", "Y"])
    save_sheet(df, output, "test_sheet")
    assert os.path.exists(output)


def test_save_sheet_correct_data(tmp_path):
    output = str(tmp_path / "out.xlsx")
    df = pd.DataFrame([["hello", "world"]], columns=["A", "B"])
    save_sheet(df, output, "data")
    wb = load_workbook(output)
    ws = wb["data"]
    assert ws.cell(1, 1).value == "A"
    assert ws.cell(2, 1).value == "hello"


def test_save_sheet_appends_without_overwriting_other_sheets(tmp_path):
    output = str(tmp_path / "out.xlsx")
    df1 = pd.DataFrame([["r1"]], columns=["C"])
    df2 = pd.DataFrame([["r2"]], columns=["D"])
    save_sheet(df1, output, "sheet1")
    save_sheet(df2, output, "sheet2")
    wb = load_workbook(output)
    assert "sheet1" in wb.sheetnames
    assert "sheet2" in wb.sheetnames


def test_save_sheet_replaces_existing_sheet(tmp_path):
    output = str(tmp_path / "out.xlsx")
    df1 = pd.DataFrame([["old"]], columns=["C"])
    df2 = pd.DataFrame([["new"]], columns=["C"])
    save_sheet(df1, output, "data")
    save_sheet(df2, output, "data")
    wb = load_workbook(output)
    ws = wb["data"]
    assert ws.cell(2, 1).value == "new"
